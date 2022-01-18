from openpyxl import Workbook, load_workbook

wb = load_workbook('PriceHistory.xlsx')
pricelst = []
onehour = []
fourhour = []
twelvehour = []
daily = []
threeday = []
print(wb.sheetnames)


def Collectph():
    for cell in ws['B']:
        pricelst.append(cell.value)
    return pricelst


ws = wb["OneHour"]
wb.active = ws
print(wb.active)
Collectph()
onehour = pricelst
print(onehour)

ws = wb["FourHour"]
wb.active = ws
pricelst = []
print(wb.active)
Collectph()
fourhour = pricelst
print(fourhour)

ws = wb["TwelveHour"]
wb.active = ws
pricelst = []
print(wb.active)
Collectph()
twelvehour = pricelst
print(twelvehour)

ws = wb["Daily"]
wb.active = ws
pricelst = []
print(wb.active)
Collectph()
daily = pricelst
print(daily)

ws = wb["ThreeDay"]
wb.active = ws
pricelst = []
print(wb.active)
Collectph()
threeday = pricelst
print(threeday)


# wb.save('PriceHistory.xlsx')
